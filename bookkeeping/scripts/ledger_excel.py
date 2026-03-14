#!/usr/bin/env -S uv run python
# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1,<4"]
# ///

from __future__ import annotations

import argparse
import calendar
import hashlib
import json
import os
import re
import tempfile
from collections import defaultdict
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Iterable, Sequence

try:
    import fcntl
except ImportError:  # pragma: no cover
    fcntl = None

from openpyxl import Workbook, load_workbook


CONFIG_FILE = "bookkeeping-config.xlsx"
LOCK_FILE = ".ledger.lock"
MONEY_QUANT = Decimal("0.01")
MONTH_WORKBOOK_RE = re.compile(r"^(?P<year>\d{4})-(?P<month>0[1-9]|1[0-2])\.xlsx$")
CONFIG_SHEETS = {
    "categories": ["direction", "level1", "level2", "keywords", "active"],
    "tags": ["tag", "active", "note"],
    "accounts": ["account", "account_type", "currency", "opening_hint", "active", "note"],
    "account_aliases": ["alias", "account", "active", "note"],
    "recurring": [
        "recurring_id", "name", "direction", "amount", "account", "from_account", "to_account",
        "level1", "level2", "tags", "note", "counterparty", "merchant", "frequency", "interval",
        "month", "day", "start_date", "end_date", "active", "last_applied_month",
    ],
    "annual_limits": ["year", "tag", "limit_amount", "note", "active"],
    "subscriptions": [
        "subscription_id", "name", "vendor", "account", "amount", "billing_cycle",
        "start_date", "renewal_date", "contract_end_date", "status", "tags", "note",
        "source_transaction_id", "source_recurring_id", "updated_at",
    ],
}
MONTHLY_SHEETS = {
    "transactions": [
        "transaction_id", "date", "direction", "amount", "account", "from_account", "to_account",
        "level1", "level2", "tags", "note", "counterparty", "merchant", "category_source",
        "recurring_id", "subscription_name", "subscription_cycle", "subscription_start_date",
        "subscription_renewal_date", "subscription_end_date", "created_at",
    ],
    "balances": ["account", "opening_balance", "closing_balance", "computed_closing", "difference", "updated_at"],
    "summary": ["metric", "value"],
}
YEARLY_KEYWORDS = ("年费", "年度", "annual", "yearly", "年订阅")
MONTHLY_KEYWORDS = ("月费", "月度", "monthly", "月订阅")

EXPENSE_CATEGORIES = {
    "餐饮": {
        "基础饮食": ["早餐", "午饭", "晚饭", "买菜", "盒饭", "食堂", "外卖", "正餐"],
        "零食饮料": ["奶茶", "咖啡", "饮料", "零食", "甜品", "水果", "酸奶"],
        "美食": ["火锅", "烧烤", "日料", "西餐", "聚餐", "自助", "餐厅"],
    },
    "天天向上": {
        "运动": ["健身", "跑步", "羽毛球", "游泳", "瑜伽", "球馆", "私教"],
        "学习资源": ["课程", "书", "教材", "训练营", "学费", "题库"],
    },
    "兴趣爱好": {
        "娱乐": ["电影", "演出", "桌游", "门票", "剧本杀"],
        "旅行": ["景点", "机票", "签证", "旅行", "旅游", "度假"],
        "摄影": ["镜头", "滤镜", "拍摄", "相机包", "云台", "修图"],
        "游戏": ["steam", "psn", "xbox", "任天堂", "游戏", "dlc", "点卡"],
    },
    "数码数字": {
        "软件买断": ["软件", "license", "买断", "授权", "app"],
        "服务订阅": ["订阅", "会员", "notion", "claude", "chatgpt", "icloud", "netflix"],
        "话费网费": ["宽带", "话费", "流量", "手机费", "网费", "通信"],
        "数码产品": ["键盘", "鼠标", "耳机", "显示器", "手机", "电脑", "ipad"],
    },
    "交通住宿": {
        "停车费": ["停车", "停车场"],
        "住宿": ["酒店", "民宿", "住宿", "宾馆"],
        "私家车费用": ["加油", "洗车", "保养", "高速", "车险", "维修"],
        "公共交通": ["地铁", "公交", "高铁", "火车", "轮渡"],
        "打车租车": ["滴滴", "uber", "出租车", "租车", "代驾"],
    },
    "服饰": {
        "鞋帽包包": ["鞋", "帽", "包", "背包", "钱包"],
        "衣裤": ["衣服", "裤子", "外套", "衬衫", "毛衣", "羽绒服"],
    },
    "医疗保健": {
        "医疗险": ["医保", "医疗险", "保险"],
        "药品费": ["药", "药店", "感冒药", "处方"],
        "保健费": ["体检", "保健", "维生素", "营养品"],
        "治疗费": ["挂号", "医院", "治疗", "口腔", "牙医", "理疗"],
    },
    "家居房产": {
        "物业管理": ["物业", "管理费"],
        "银行手续": ["手续费", "银行费", "账户管理费"],
        "房贷": ["房贷", "按揭"],
        "日常用品": ["纸巾", "清洁", "厨房用品", "洗衣液", "家庭用品"],
        "家具家电": ["冰箱", "空调", "床", "桌子", "椅子", "家电"],
        "水电煤气": ["水费", "电费", "燃气", "煤气"],
    },
    "其他": {
        "账户修正": ["修正", "差错", "调账"],
        "杂项": ["杂项", "misc"],
        "偶然支出": ["临时", "意外", "偶发"],
        "爱情": ["约会", "恋爱", "纪念日"],
        "礼物": ["礼物", "花", "鲜花", "礼盒"],
        "家庭杂项": ["家庭", "家用", "父母", "亲属"],
    },
    "儿子": {
        "陪伴": ["乐园", "亲子", "陪伴"],
        "教育": ["培训", "作业", "课程", "学校", "学费"],
        "养育": ["奶粉", "尿不湿", "玩具", "儿童医院", "辅食"],
        "剩余": ["儿子", "孩子", "宝宝"],
    },
    "人情往来": {
        "送礼请客": ["红包", "送礼", "请客", "礼金", "随礼"],
        "交际聚会": ["聚会", "社交", "聚餐", "团建"],
    },
}
INCOME_CATEGORIES = {
    "收入": {
        "工资": ["工资", "salary", "发薪", "薪资"],
        "奖金": ["奖金", "绩效", "bonus"],
        "报销": ["报销", "reimbursement"],
        "理财": ["分红", "理财", "利息", "基金", "股票"],
        "红包礼金": ["红包", "礼金", "礼物"],
        "其他收入": ["退款", "返现", "闲置", "收款", "其他收入"],
    }
}
DEFAULT_ACCOUNTS = [
    ("现金", "cash", "CNY", "0", True, "默认现金账户"),
    ("银行卡", "bank", "CNY", "0", True, "默认银行卡账户"),
    ("支付宝", "wallet", "CNY", "0", True, "默认支付宝账户"),
    ("微信", "wallet", "CNY", "0", True, "默认微信账户"),
]


@dataclass
class CategoryMatch:
    level1: str
    level2: str
    score: int
    reason: str


@dataclass
class TransactionRef:
    workbook_path: Path
    month: str
    row_index: int
    data: dict[str, str]


def parse_money(raw: str | int | float | Decimal) -> Decimal:
    try:
        value = Decimal(str(raw)).quantize(MONEY_QUANT, rounding=ROUND_HALF_UP)
    except InvalidOperation as exc:
        raise ValueError(f"invalid money amount: {raw}") from exc
    if value < 0:
        raise ValueError("amount must be non-negative")
    return value


def now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def parse_date(raw: str) -> date:
    return datetime.strptime(raw, "%Y-%m-%d").date()


def month_string(day: date) -> str:
    return f"{day.year:04d}-{day.month:02d}"


def parse_month(raw: str) -> tuple[int, int]:
    year, month = raw.split("-", 1)
    return int(year), int(month)


def month_path(root: Path, month: str) -> Path:
    year, _ = parse_month(month)
    return root / f"{year:04d}" / f"{month}.xlsx"


def config_path(root: Path) -> Path:
    return root / CONFIG_FILE


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def save_workbook_atomic(workbook: Workbook, path: Path) -> None:
    ensure_dir(path.parent)
    fd, temp_name = tempfile.mkstemp(prefix=f".{path.stem}.", suffix=".tmp", dir=str(path.parent))
    os.close(fd)
    temp_path = Path(temp_name)
    try:
        workbook.save(temp_path)
        os.replace(temp_path, path)
    finally:
        if temp_path.exists():
            temp_path.unlink()


@contextmanager
def ledger_lock(root: Path):
    ensure_dir(root)
    lock_file = root / LOCK_FILE
    with lock_file.open("a+", encoding="utf-8") as handle:
        if fcntl is not None:
            try:
                fcntl.flock(handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
            except BlockingIOError as exc:
                raise SystemExit(f"ledger is busy: {root} (lock file: {lock_file})") from exc
        handle.seek(0)
        handle.truncate()
        handle.write(f"pid={os.getpid()} acquired_at={now_iso()}\n")
        handle.flush()
        try:
            yield
        finally:
            if fcntl is not None:
                fcntl.flock(handle.fileno(), fcntl.LOCK_UN)


def ensure_headers(sheet, headers: Sequence[str]) -> bool:
    changed = False
    if sheet.max_row == 1 and sheet["A1"].value is None:
        changed = True
    existing = [sheet.cell(row=1, column=i).value for i in range(1, len(headers) + 1)]
    if existing != list(headers):
        changed = True
        for idx, name in enumerate(headers, start=1):
            sheet.cell(row=1, column=idx, value=name)
    return changed


def rows_as_dicts(sheet) -> list[dict[str, str]]:
    headers = ["" if cell.value is None else str(cell.value) for cell in sheet[1]]
    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if all(value in (None, "") for value in values):
            continue
        row = {}
        for key, value in zip(headers, values):
            row[key] = "" if value is None else str(value)
        rows.append(row)
    return rows


def indexed_rows_as_dicts(sheet) -> list[tuple[int, dict[str, str]]]:
    headers = ["" if cell.value is None else str(cell.value) for cell in sheet[1]]
    rows = []
    for row_idx, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(value in (None, "") for value in values):
            continue
        row = {}
        for key, value in zip(headers, values):
            row[key] = "" if value is None else str(value)
        rows.append((row_idx, row))
    return rows


def append_row(sheet, headers: Sequence[str], values: dict[str, object]) -> None:
    ensure_headers(sheet, headers)
    sheet.append([values.get(header, "") for header in headers])


def truthy(value: str) -> bool:
    return str(value).strip().lower() in {"1", "true", "yes", "y"}


def previous_month(month: str) -> str:
    year, mm = parse_month(month)
    return f"{year - 1:04d}-12" if mm == 1 else f"{year:04d}-{mm - 1:02d}"


def normalize_tags(tags: Iterable[str]) -> str:
    seen = set()
    result = []
    for tag in tags:
        clean = tag.strip()
        if not clean or clean in seen:
            continue
        seen.add(clean)
        result.append(clean)
    return ",".join(result)


def split_tags(raw_tags: str) -> list[str]:
    return [item.strip() for item in raw_tags.split(",") if item and item.strip()]


def transaction_id(prefix: str, month: str, parts: Sequence[str]) -> str:
    digest = hashlib.md5("|".join(parts).encode("utf-8")).hexdigest()[:8]
    return f"{prefix}-{month.replace('-', '')}-{digest}"


def stable_id(prefix: str, parts: Sequence[str]) -> str:
    digest = hashlib.md5("|".join(parts).encode("utf-8")).hexdigest()[:10]
    return f"{prefix}-{digest}"


def combined_text(parts: Iterable[str]) -> str:
    return " ".join(part.strip().lower() for part in parts if part and part.strip())


def seed_config_workbook(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    categories = wb.create_sheet("categories")
    ensure_headers(categories, CONFIG_SHEETS["categories"])
    for level1, children in EXPENSE_CATEGORIES.items():
        for level2, keywords in children.items():
            categories.append(["expense", level1, level2, ",".join(keywords), True])
    for level1, children in INCOME_CATEGORIES.items():
        for level2, keywords in children.items():
            categories.append(["income", level1, level2, ",".join(keywords), True])

    tags = wb.create_sheet("tags")
    ensure_headers(tags, CONFIG_SHEETS["tags"])

    accounts = wb.create_sheet("accounts")
    ensure_headers(accounts, CONFIG_SHEETS["accounts"])
    for row in DEFAULT_ACCOUNTS:
        accounts.append(list(row))

    account_aliases = wb.create_sheet("account_aliases")
    ensure_headers(account_aliases, CONFIG_SHEETS["account_aliases"])

    recurring = wb.create_sheet("recurring")
    ensure_headers(recurring, CONFIG_SHEETS["recurring"])

    annual_limits = wb.create_sheet("annual_limits")
    ensure_headers(annual_limits, CONFIG_SHEETS["annual_limits"])

    subscriptions = wb.create_sheet("subscriptions")
    ensure_headers(subscriptions, CONFIG_SHEETS["subscriptions"])

    ensure_dir(path.parent)
    save_workbook_atomic(wb, path)


def ensure_config_workbook(root: Path) -> Path:
    path = config_path(root)
    if not path.exists():
        seed_config_workbook(path)
        return path
    wb = load_workbook(path)
    changed = False
    for sheet_name, headers in CONFIG_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
            changed = True
        changed = ensure_headers(wb[sheet_name], headers) or changed
    if changed:
        save_workbook_atomic(wb, path)
    return path


def load_config_rows(root: Path, sheet_name: str) -> list[dict[str, str]]:
    ensure_config_workbook(root)
    wb = load_workbook(config_path(root))
    return rows_as_dicts(wb[sheet_name])


def active_accounts(root: Path) -> list[str]:
    return [row["account"] for row in load_config_rows(root, "accounts") if truthy(row.get("active", "true"))]


def account_alias_map(root: Path) -> dict[str, str]:
    active = set(active_accounts(root))
    mapping: dict[str, str] = {}
    for row in load_config_rows(root, "account_aliases"):
        alias = row.get("alias", "").strip()
        account = row.get("account", "").strip()
        if not alias or not account:
            continue
        if not truthy(row.get("active", "true")):
            continue
        if account not in active:
            continue
        mapping[alias] = account
    return mapping


def resolve_account(root: Path, account: str) -> str | None:
    active = set(active_accounts(root))
    if account in active:
        return account
    mapping = account_alias_map(root)
    return mapping.get(account)


def ensure_balance_row(sheet, account: str, opening_hint: str = "0.00") -> int:
    for row_idx in range(2, sheet.max_row + 1):
        if sheet.cell(row=row_idx, column=1).value == account:
            return row_idx
    sheet.append([account, opening_hint, "", "0.00", "", now_iso()])
    return sheet.max_row


def load_previous_closing(root: Path, month: str, account: str) -> str:
    prev = month_path(root, previous_month(month))
    if not prev.exists():
        return "0.00"
    wb = load_workbook(prev)
    if "balances" not in wb.sheetnames:
        return "0.00"
    for row in rows_as_dicts(wb["balances"]):
        if row["account"] == account and row.get("closing_balance"):
            return row["closing_balance"]
    return "0.00"


def ensure_month_workbook(root: Path, month: str) -> Path:
    ensure_config_workbook(root)
    path = month_path(root, month)
    if not path.exists():
        ensure_dir(path.parent)
        wb = Workbook()
        tx = wb.active
        tx.title = "transactions"
        ensure_headers(tx, MONTHLY_SHEETS["transactions"])
        balances = wb.create_sheet("balances")
        ensure_headers(balances, MONTHLY_SHEETS["balances"])
        for account in active_accounts(root):
            balances.append([account, load_previous_closing(root, month, account), "", "0.00", "", now_iso()])
        summary = wb.create_sheet("summary")
        ensure_headers(summary, MONTHLY_SHEETS["summary"])
        save_workbook_atomic(wb, path)
        return path

    wb = load_workbook(path)
    changed = False
    for sheet_name, headers in MONTHLY_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
            changed = True
        changed = ensure_headers(wb[sheet_name], headers) or changed
    balances_sheet = wb["balances"]
    for account in active_accounts(root):
        before_row = balances_sheet.max_row
        ensure_balance_row(balances_sheet, account, load_previous_closing(root, month, account))
        if balances_sheet.max_row != before_row:
            changed = True
    if changed:
        save_workbook_atomic(wb, path)
    return path


def bootstrap(root: Path, month: str | None) -> list[str]:
    ensure_dir(root)
    actions = []
    if not config_path(root).exists():
        ensure_config_workbook(root)
        actions.append(f"created {config_path(root)}")
    else:
        ensure_config_workbook(root)
        actions.append(f"kept {config_path(root)}")
    target_month = month or month_string(date.today())
    actions.append(f"ready {ensure_month_workbook(root, target_month)}")
    return actions


def classify_category(root: Path, direction: str, note: str, merchant: str, counterparty: str, tags: str) -> tuple[CategoryMatch | None, list[CategoryMatch]]:
    text = combined_text([note, merchant, counterparty, tags])
    if not text:
        return None, []
    candidates = []
    for row in load_config_rows(root, "categories"):
        if not truthy(row.get("active", "true")) or row["direction"] not in {direction, "both"}:
            continue
        keywords = [item.strip().lower() for item in row.get("keywords", "").split(",") if item.strip()]
        score = 0
        hits = []
        for keyword in keywords:
            if keyword in text:
                score += 2
                hits.append(keyword)
        if row["level1"].lower() in text:
            score += 1
            hits.append(row["level1"])
        if row["level2"].lower() in text:
            score += 2
            hits.append(row["level2"])
        if score > 0:
            candidates.append(CategoryMatch(row["level1"], row["level2"], score, ",".join(hits)))
    candidates.sort(key=lambda item: (-item.score, item.level1, item.level2))
    best = candidates[0] if candidates else None
    if best and len(candidates) > 1 and best.score - candidates[1].score <= 1:
        return None, candidates[:3]
    return best, candidates[:3]


def validate_category(root: Path, direction: str, level1: str, level2: str) -> None:
    for row in load_config_rows(root, "categories"):
        if truthy(row.get("active", "true")) and row["direction"] in {direction, "both"} and row["level1"] == level1 and row["level2"] == level2:
            return
    raise ValueError(f"unknown category: {direction}/{level1}/{level2}")


def validate_account(root: Path, account: str) -> str:
    resolved = resolve_account(root, account)
    if not resolved:
        raise ValueError(f"unknown account: {account}")
    return resolved


def rewrite_summary_sheet(summary_sheet, transactions: list[dict[str, str]], balances_sheet, root: Path, month: str) -> None:
    for row_idx in range(summary_sheet.max_row, 1, -1):
        summary_sheet.delete_rows(row_idx)
    income_total = Decimal("0.00")
    expense_total = Decimal("0.00")
    per_level1 = defaultdict(lambda: Decimal("0.00"))
    for row in transactions:
        amount = parse_money(row["amount"])
        if row["direction"] == "income":
            income_total += amount
        elif row["direction"] == "expense":
            expense_total += amount
            per_level1[row["level1"] or "未分类"] += amount
    append_row(summary_sheet, MONTHLY_SHEETS["summary"], {"metric": "month", "value": month})
    append_row(summary_sheet, MONTHLY_SHEETS["summary"], {"metric": "income_total", "value": str(income_total.quantize(MONEY_QUANT))})
    append_row(summary_sheet, MONTHLY_SHEETS["summary"], {"metric": "expense_total", "value": str(expense_total.quantize(MONEY_QUANT))})
    append_row(summary_sheet, MONTHLY_SHEETS["summary"], {"metric": "net_cashflow", "value": str((income_total - expense_total).quantize(MONEY_QUANT))})
    for level1, total in sorted(per_level1.items()):
        append_row(summary_sheet, MONTHLY_SHEETS["summary"], {"metric": f"expense:{level1}", "value": str(total.quantize(MONEY_QUANT))})
    for row in rows_as_dicts(balances_sheet):
        append_row(summary_sheet, MONTHLY_SHEETS["summary"], {"metric": f"balance:{row['account']}", "value": json.dumps(row, ensure_ascii=False)})
    for item in collect_limit_usage(root, int(month[:4])):
        append_row(summary_sheet, MONTHLY_SHEETS["summary"], {"metric": f"limit:{item['tag']}", "value": json.dumps(item, ensure_ascii=False)})


def update_balances(root: Path, month: str) -> None:
    ensure_month_workbook(root, month)
    wb = load_workbook(month_path(root, month))
    transactions = rows_as_dicts(wb["transactions"])
    balances_sheet = wb["balances"]
    per_account = {account: Decimal("0.00") for account in active_accounts(root)}
    for row in transactions:
        amount = parse_money(row["amount"])
        if row["direction"] == "expense":
            per_account[row["account"]] -= amount
        elif row["direction"] == "income":
            per_account[row["account"]] += amount
        elif row["direction"] == "transfer":
            per_account[row["from_account"]] -= amount
            per_account[row["to_account"]] += amount
    for account in active_accounts(root):
        row_idx = ensure_balance_row(balances_sheet, account, load_previous_closing(root, month, account))
        opening = parse_money(balances_sheet.cell(row=row_idx, column=2).value or "0")
        computed = (opening + per_account[account]).quantize(MONEY_QUANT)
        closing_raw = balances_sheet.cell(row=row_idx, column=3).value
        balances_sheet.cell(row=row_idx, column=4, value=str(computed))
        if closing_raw not in (None, ""):
            closing = parse_money(closing_raw)
            balances_sheet.cell(row=row_idx, column=5, value=str((closing - computed).quantize(MONEY_QUANT)))
        else:
            balances_sheet.cell(row=row_idx, column=5, value="")
        balances_sheet.cell(row=row_idx, column=6, value=now_iso())
    rewrite_summary_sheet(wb["summary"], transactions, balances_sheet, root, month)
    save_workbook_atomic(wb, month_path(root, month))


def infer_subscription_cycle(*parts: str, fallback: str = "") -> str:
    text = combined_text(parts)
    if fallback in {"monthly", "yearly"}:
        return fallback
    if any(keyword in text for keyword in YEARLY_KEYWORDS):
        return "yearly"
    if any(keyword in text for keyword in MONTHLY_KEYWORDS):
        return "monthly"
    return ""


def shift_cycle(raw_date: str, cycle: str, interval: int = 1) -> str:
    start = parse_date(raw_date)
    if cycle == "monthly":
        months = interval
    elif cycle == "yearly":
        months = 12 * interval
    else:
        return ""
    year = start.year + (start.month - 1 + months) // 12
    month = (start.month - 1 + months) % 12 + 1
    day = min(start.day, calendar.monthrange(year, month)[1])
    return date(year, month, day).isoformat()


def is_service_subscription(row: dict[str, str]) -> bool:
    return row.get("direction") == "expense" and row.get("level2") == "服务订阅"


def infer_subscription_name(row: dict[str, str], fallback_name: str = "") -> str:
    return (
        row.get("subscription_name")
        or fallback_name
        or row.get("merchant")
        or row.get("counterparty")
        or row.get("note")
        or row.get("level2")
    )


def upsert_subscription_item(store: dict[str, dict[str, str]], item: dict[str, str], event_date: str) -> None:
    existing = store.get(item["subscription_id"])
    if existing is None:
        store[item["subscription_id"]] = item
        return
    existing_date = existing.get("_event_date", existing.get("start_date", ""))
    if item.get("start_date") and (not existing.get("start_date") or item["start_date"] < existing["start_date"]):
        existing["start_date"] = item["start_date"]
    if item.get("renewal_date") and (not existing.get("renewal_date") or item["renewal_date"] > existing["renewal_date"]):
        existing["renewal_date"] = item["renewal_date"]
    if item.get("contract_end_date") and (not existing.get("contract_end_date") or item["contract_end_date"] > existing["contract_end_date"]):
        existing["contract_end_date"] = item["contract_end_date"]
    existing["tags"] = normalize_tags(split_tags(existing.get("tags", "")) + split_tags(item.get("tags", "")))
    if event_date >= existing_date:
        for key in ("name", "vendor", "account", "amount", "billing_cycle", "note", "source_transaction_id", "source_recurring_id"):
            if item.get(key):
                existing[key] = item[key]
        existing["_event_date"] = event_date


def subscription_status(renewal_date: str, contract_end_date: str) -> str:
    today = date.today().isoformat()
    if contract_end_date and contract_end_date < today:
        return "expired"
    if renewal_date and renewal_date < today:
        return "due"
    return "active"


def rebuild_subscriptions(root: Path) -> None:
    ensure_config_workbook(root)
    items: dict[str, dict[str, str]] = {}

    for row in load_config_rows(root, "recurring"):
        if not truthy(row.get("active", "true")) or row.get("direction") != "expense" or row.get("level2") != "服务订阅":
            continue
        billing_cycle = row.get("frequency") if row.get("frequency") in {"monthly", "yearly"} else infer_subscription_cycle(row.get("name", ""), row.get("note", ""), row.get("merchant", ""))
        start_date = row.get("start_date", "")
        subscription_id = stable_id("SB", [row.get("recurring_id", ""), row.get("name", ""), row.get("account", "")])
        item = {
            "subscription_id": subscription_id,
            "name": infer_subscription_name(row, row.get("name", "")),
            "vendor": row.get("merchant", "") or row.get("counterparty", ""),
            "account": row.get("account", ""),
            "amount": str(parse_money(row["amount"])),
            "billing_cycle": billing_cycle,
            "start_date": start_date,
            "renewal_date": shift_cycle(start_date, billing_cycle, int(row.get("interval") or "1")) if start_date else "",
            "contract_end_date": row.get("end_date", ""),
            "status": "",
            "tags": normalize_tags(split_tags(row.get("tags", ""))),
            "note": row.get("note", "") or row.get("name", ""),
            "source_transaction_id": "",
            "source_recurring_id": row.get("recurring_id", ""),
            "updated_at": "",
            "_event_date": start_date,
        }
        upsert_subscription_item(items, item, start_date)

    for ref in load_transaction_refs(root):
        row = ref.data
        if not is_service_subscription(row):
            continue
        billing_cycle = row.get("subscription_cycle") or infer_subscription_cycle(
            row.get("subscription_name", ""),
            row.get("note", ""),
            row.get("merchant", ""),
            row.get("counterparty", ""),
            row.get("tags", ""),
        )
        name = infer_subscription_name(row)
        key_parts = [row.get("recurring_id", ""), name, row.get("merchant", ""), row.get("account", ""), billing_cycle]
        subscription_id = stable_id("SB", key_parts)
        start_date = row.get("subscription_start_date") or row.get("date", "")
        renewal_date = row.get("subscription_renewal_date") or shift_cycle(start_date, billing_cycle)
        contract_end_date = row.get("subscription_end_date", "")
        item = {
            "subscription_id": subscription_id,
            "name": name,
            "vendor": row.get("merchant", "") or row.get("counterparty", ""),
            "account": row.get("account", ""),
            "amount": str(parse_money(row["amount"])),
            "billing_cycle": billing_cycle,
            "start_date": start_date,
            "renewal_date": renewal_date,
            "contract_end_date": contract_end_date,
            "status": "",
            "tags": normalize_tags(split_tags(row.get("tags", ""))),
            "note": row.get("note", ""),
            "source_transaction_id": row.get("transaction_id", ""),
            "source_recurring_id": row.get("recurring_id", ""),
            "updated_at": "",
            "_event_date": row.get("date", ""),
        }
        upsert_subscription_item(items, item, row.get("date", ""))

    wb = load_workbook(config_path(root))
    sheet = wb["subscriptions"]
    for row_idx in range(sheet.max_row, 1, -1):
        sheet.delete_rows(row_idx)
    final_rows = []
    for item in items.values():
        item["status"] = subscription_status(item.get("renewal_date", ""), item.get("contract_end_date", ""))
        item["updated_at"] = now_iso()
        final_rows.append(item)
    final_rows.sort(key=lambda item: (item.get("renewal_date") or "9999-12-31", item["name"]))
    for item in final_rows:
        append_row(sheet, CONFIG_SHEETS["subscriptions"], item)
    save_workbook_atomic(wb, config_path(root))


def add_transaction(
    root: Path,
    raw_date: str,
    direction: str,
    amount: Decimal,
    account: str,
    level1: str,
    level2: str,
    tags: str,
    note: str,
    counterparty: str,
    merchant: str,
    recurring_id: str = "",
    subscription_name: str = "",
    subscription_cycle: str = "",
    subscription_start_date: str = "",
    subscription_renewal_date: str = "",
    subscription_end_date: str = "",
) -> dict[str, str]:
    month = month_string(parse_date(raw_date))
    ensure_month_workbook(root, month)
    account = validate_account(root, account)
    category_source = "manual"
    suggestions = []
    if not (level1 and level2):
        match, suggestions = classify_category(root, direction, note, merchant, counterparty, tags)
        if match:
            level1, level2 = match.level1, match.level2
            category_source = "auto"
        else:
            level1, level2 = level1 or "", level2 or ""
            category_source = "unclassified"
    else:
        validate_category(root, direction, level1, level2)

    record = {
        "transaction_id": transaction_id("TX", month, [raw_date, direction, str(amount), account, note, tags, recurring_id]),
        "date": raw_date,
        "direction": direction,
        "amount": str(amount),
        "account": account,
        "from_account": "",
        "to_account": "",
        "level1": level1,
        "level2": level2,
        "tags": normalize_tags(split_tags(tags)),
        "note": note,
        "counterparty": counterparty,
        "merchant": merchant,
        "category_source": category_source,
        "recurring_id": recurring_id,
        "subscription_name": subscription_name,
        "subscription_cycle": subscription_cycle,
        "subscription_start_date": subscription_start_date,
        "subscription_renewal_date": subscription_renewal_date,
        "subscription_end_date": subscription_end_date,
        "created_at": now_iso(),
    }
    wb = load_workbook(month_path(root, month))
    append_row(wb["transactions"], MONTHLY_SHEETS["transactions"], record)
    save_workbook_atomic(wb, month_path(root, month))
    update_balances(root, month)
    rebuild_subscriptions(root)
    if category_source == "unclassified" and suggestions:
        record["suggestion"] = " | ".join(f"{item.level1}/{item.level2}:{item.score}" for item in suggestions)
    elif category_source == "unclassified":
        record["suggestion"] = "建议补充新分类或为现有分类增加关键词"
    return record


def add_transfer(root: Path, raw_date: str, amount: Decimal, from_account: str, to_account: str, tags: str, note: str) -> dict[str, str]:
    month = month_string(parse_date(raw_date))
    ensure_month_workbook(root, month)
    from_account = validate_account(root, from_account)
    to_account = validate_account(root, to_account)
    record = {
        "transaction_id": transaction_id("TR", month, [raw_date, str(amount), from_account, to_account, note]),
        "date": raw_date,
        "direction": "transfer",
        "amount": str(amount),
        "account": "",
        "from_account": from_account,
        "to_account": to_account,
        "level1": "",
        "level2": "",
        "tags": normalize_tags(split_tags(tags)),
        "note": note,
        "counterparty": "",
        "merchant": "",
        "category_source": "manual",
        "recurring_id": "",
        "subscription_name": "",
        "subscription_cycle": "",
        "subscription_start_date": "",
        "subscription_renewal_date": "",
        "subscription_end_date": "",
        "created_at": now_iso(),
    }
    wb = load_workbook(month_path(root, month))
    append_row(wb["transactions"], MONTHLY_SHEETS["transactions"], record)
    save_workbook_atomic(wb, month_path(root, month))
    update_balances(root, month)
    rebuild_subscriptions(root)
    return record


def set_balance(root: Path, month: str, account: str, amount: Decimal, column: str) -> dict[str, str]:
    ensure_month_workbook(root, month)
    account = validate_account(root, account)
    wb = load_workbook(month_path(root, month))
    sheet = wb["balances"]
    row_idx = ensure_balance_row(sheet, account, load_previous_closing(root, month, account))
    col_idx = MONTHLY_SHEETS["balances"].index(column) + 1
    sheet.cell(row=row_idx, column=col_idx, value=str(amount))
    sheet.cell(row=row_idx, column=6, value=now_iso())
    save_workbook_atomic(wb, month_path(root, month))
    update_balances(root, month)
    wb = load_workbook(month_path(root, month))
    for row in rows_as_dicts(wb["balances"]):
        if row["account"] == account:
            return row
    raise RuntimeError("balance row missing")


def list_accounts(root: Path, include_inactive: bool) -> dict[str, object]:
    ensure_config_workbook(root)
    account_rows = load_config_rows(root, "accounts")
    if not include_inactive:
        account_rows = [row for row in account_rows if truthy(row.get("active", "true"))]
    aliases = load_config_rows(root, "account_aliases")
    account_names = {row["account"] for row in account_rows}
    alias_map: dict[str, str] = {}
    for row in aliases:
        alias = row.get("alias", "").strip()
        account = row.get("account", "").strip()
        if not alias or not account:
            continue
        if not include_inactive and not truthy(row.get("active", "true")):
            continue
        if account not in account_names:
            continue
        alias_map[alias] = account
    return {
        "count": len(account_rows),
        "accounts": account_rows,
        "alias_count": len(alias_map),
        "alias_map": alias_map,
    }


def upsert_account(
    root: Path,
    account: str,
    account_type: str,
    currency: str,
    opening_hint: str,
    active: bool,
    note: str,
    aliases: list[str],
) -> dict[str, object]:
    ensure_config_workbook(root)
    wb = load_workbook(config_path(root))
    account_sheet = wb["accounts"]
    canonical_opening = str(parse_money(opening_hint))
    active_value = "True" if active else "False"
    updated = False
    for row_idx in range(2, account_sheet.max_row + 1):
        if str(account_sheet.cell(row=row_idx, column=1).value or "") == account:
            account_sheet.cell(row=row_idx, column=2, value=account_type)
            account_sheet.cell(row=row_idx, column=3, value=currency)
            account_sheet.cell(row=row_idx, column=4, value=canonical_opening)
            account_sheet.cell(row=row_idx, column=5, value=active_value)
            account_sheet.cell(row=row_idx, column=6, value=note)
            updated = True
            break
    if not updated:
        account_sheet.append([account, account_type, currency, canonical_opening, active_value, note])

    alias_sheet = wb["account_aliases"]
    applied_aliases: list[str] = []
    for raw_alias in aliases:
        alias = raw_alias.strip()
        if not alias or alias == account:
            continue
        found = False
        for row_idx in range(2, alias_sheet.max_row + 1):
            if str(alias_sheet.cell(row=row_idx, column=1).value or "") == alias:
                alias_sheet.cell(row=row_idx, column=2, value=account)
                alias_sheet.cell(row=row_idx, column=3, value="True")
                alias_sheet.cell(row=row_idx, column=4, value=f"alias for {account}")
                found = True
                break
        if not found:
            alias_sheet.append([alias, account, "True", f"alias for {account}"])
        applied_aliases.append(alias)

    save_workbook_atomic(wb, config_path(root))
    return {
        "account": account,
        "account_type": account_type,
        "currency": currency,
        "opening_hint": canonical_opening,
        "active": active,
        "note": note,
        "aliases": applied_aliases,
        "updated": updated,
    }


def upsert_limit(root: Path, year: int, tag: str, amount: Decimal, note: str) -> dict[str, str]:
    ensure_config_workbook(root)
    wb = load_workbook(config_path(root))
    sheet = wb["annual_limits"]
    updated = False
    for row_idx in range(2, sheet.max_row + 1):
        if sheet.cell(row=row_idx, column=1).value == year and sheet.cell(row=row_idx, column=2).value == tag:
            sheet.cell(row=row_idx, column=3, value=str(amount))
            sheet.cell(row=row_idx, column=4, value=note)
            sheet.cell(row=row_idx, column=5, value=True)
            updated = True
            break
    if not updated:
        sheet.append([year, tag, str(amount), note, True])
    save_workbook_atomic(wb, config_path(root))
    return {"year": str(year), "tag": tag, "limit_amount": str(amount), "note": note}


def add_recurring(
    root: Path,
    name: str,
    direction: str,
    amount: Decimal,
    account: str,
    from_account: str,
    to_account: str,
    level1: str,
    level2: str,
    tags: str,
    note: str,
    counterparty: str,
    merchant: str,
    frequency: str,
    interval: int,
    month_value: int,
    day: int,
    start_date: str,
    end_date: str,
) -> dict[str, str]:
    if frequency not in {"monthly", "yearly"}:
        raise ValueError("frequency must be monthly or yearly")
    if direction in {"expense", "income"}:
        account = validate_account(root, account)
        validate_category(root, direction, level1, level2)
    if direction == "transfer":
        from_account = validate_account(root, from_account)
        to_account = validate_account(root, to_account)
    ensure_config_workbook(root)
    wb = load_workbook(config_path(root))
    sheet = wb["recurring"]
    record = {
        "recurring_id": transaction_id("RC", month_string(parse_date(start_date)), [name, direction, str(amount), start_date]),
        "name": name,
        "direction": direction,
        "amount": str(amount),
        "account": account,
        "from_account": from_account,
        "to_account": to_account,
        "level1": level1,
        "level2": level2,
        "tags": normalize_tags(split_tags(tags)),
        "note": note,
        "counterparty": counterparty,
        "merchant": merchant,
        "frequency": frequency,
        "interval": str(interval),
        "month": str(month_value or ""),
        "day": str(day),
        "start_date": start_date,
        "end_date": end_date,
        "active": "True",
        "last_applied_month": "",
    }
    append_row(sheet, CONFIG_SHEETS["recurring"], record)
    save_workbook_atomic(wb, config_path(root))
    rebuild_subscriptions(root)
    return record


def recurring_due(row: dict[str, str], target_month: str) -> bool:
    start = parse_date(row["start_date"])
    target_year, target_mm = parse_month(target_month)
    if (target_year, target_mm) < (start.year, start.month):
        return False
    if row.get("end_date"):
        end = parse_date(row["end_date"])
        if date(target_year, target_mm, 1) > date(end.year, end.month, 1):
            return False
    interval = int(row.get("interval") or "1")
    if row["frequency"] == "monthly":
        delta = (target_year - start.year) * 12 + (target_mm - start.month)
        return delta % interval == 0
    if row["frequency"] == "yearly":
        return int(row.get("month") or "0") == target_mm and (target_year - start.year) % interval == 0
    return False


def due_day(target_month: str, raw_day: str) -> str:
    year, mm = parse_month(target_month)
    target_day = min(int(raw_day or "1"), calendar.monthrange(year, mm)[1])
    return f"{target_month}-{target_day:02d}"


def apply_recurring(root: Path, month: str) -> list[dict[str, str]]:
    ensure_month_workbook(root, month)
    cfg = load_workbook(config_path(root))
    sheet = cfg["recurring"]
    applied = []
    for row_idx in range(2, sheet.max_row + 1):
        row = {
            header: "" if sheet.cell(row=row_idx, column=i + 1).value is None else str(sheet.cell(row=row_idx, column=i + 1).value)
            for i, header in enumerate(CONFIG_SHEETS["recurring"])
        }
        if not truthy(row.get("active", "true")) or row.get("last_applied_month") == month or not recurring_due(row, month):
            continue
        tx_date = due_day(month, row.get("day") or "1")
        if row["direction"] == "transfer":
            result = add_transfer(root, tx_date, parse_money(row["amount"]), row["from_account"], row["to_account"], row["tags"], row["note"] or row["name"])
        else:
            subscription_cycle = row["frequency"] if row.get("level2") == "服务订阅" else ""
            result = add_transaction(
                root,
                tx_date,
                row["direction"],
                parse_money(row["amount"]),
                row["account"],
                row["level1"],
                row["level2"],
                row["tags"],
                row["note"] or row["name"],
                row["counterparty"],
                row["merchant"],
                row["recurring_id"],
                row["name"] if row.get("level2") == "服务订阅" else "",
                subscription_cycle,
                row["start_date"] if row.get("level2") == "服务订阅" else "",
                shift_cycle(tx_date, subscription_cycle, int(row.get("interval") or "1")) if row.get("level2") == "服务订阅" else "",
                row.get("end_date", "") if row.get("level2") == "服务订阅" else "",
            )
        sheet.cell(row=row_idx, column=21, value=month)
        applied.append(result)
    save_workbook_atomic(cfg, config_path(root))
    rebuild_subscriptions(root)
    return applied


def collect_limit_usage(root: Path, year: int) -> list[dict[str, str]]:
    limits = {(int(row["year"]), row["tag"]): parse_money(row["limit_amount"]) for row in load_config_rows(root, "annual_limits") if truthy(row.get("active", "true"))}
    usage = defaultdict(lambda: Decimal("0.00"))
    year_dir = root / f"{year:04d}"
    if year_dir.exists():
        for workbook_path in sorted(year_dir.glob("*.xlsx")):
            match = MONTH_WORKBOOK_RE.match(workbook_path.name)
            if not match or match.group("year") != f"{year:04d}":
                continue
            try:
                wb = load_workbook(workbook_path)
            except Exception:
                continue
            if "transactions" not in wb.sheetnames:
                continue
            for row in rows_as_dicts(wb["transactions"]):
                if row["direction"] != "expense":
                    continue
                amount = parse_money(row["amount"])
                for tag in split_tags(row["tags"]):
                    if (year, tag) in limits:
                        usage[tag] += amount
    report = []
    for (limit_year, tag), limit_amount in sorted(limits.items()):
        spent = usage[tag].quantize(MONEY_QUANT)
        remaining = (limit_amount - spent).quantize(MONEY_QUANT)
        report.append({
            "year": str(limit_year),
            "tag": tag,
            "limit_amount": str(limit_amount),
            "spent_amount": str(spent),
            "remaining_amount": str(remaining),
            "status": "exceeded" if remaining < 0 else "ok",
        })
    return report


def month_report(root: Path, month: str) -> dict[str, object]:
    ensure_month_workbook(root, month)
    update_balances(root, month)
    wb = load_workbook(month_path(root, month))
    return {
        "month": month,
        "transaction_count": len(rows_as_dicts(wb["transactions"])),
        "balances": rows_as_dicts(wb["balances"]),
        "summary": rows_as_dicts(wb["summary"]),
    }


def workbook_open_error(path: Path) -> str:
    if not path.exists():
        return ""
    try:
        load_workbook(path)
        return ""
    except Exception as exc:  # pragma: no cover
        return f"{type(exc).__name__}: {exc}"


def corrupted_backup_path(path: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d%H%M%S")
    return path.with_name(f"{path.stem}.corrupted-{stamp}.badzip")


def repair_month(root: Path, month: str, execute: bool) -> dict[str, object]:
    ensure_config_workbook(root)
    target = month_path(root, month)
    if not target.exists():
        if execute:
            ensure_month_workbook(root, month)
            update_balances(root, month)
            return {
                "month": month,
                "target_path": str(target),
                "is_corrupted": False,
                "error": "",
                "action": "recreated-missing",
                "backup_path": "",
                "rebuilt_path": str(target),
            }
        return {
            "month": month,
            "target_path": str(target),
            "is_corrupted": False,
            "error": "missing workbook",
            "action": "dry-run-missing",
            "backup_path": "",
            "rebuilt_path": "",
        }

    error = workbook_open_error(target)
    if not error:
        return {
            "month": month,
            "target_path": str(target),
            "is_corrupted": False,
            "error": "",
            "action": "healthy",
            "backup_path": "",
            "rebuilt_path": "",
        }

    backup_path = corrupted_backup_path(target)
    if not execute:
        return {
            "month": month,
            "target_path": str(target),
            "is_corrupted": True,
            "error": error,
            "action": "dry-run",
            "backup_path": str(backup_path),
            "rebuilt_path": "",
        }

    target.rename(backup_path)
    ensure_month_workbook(root, month)
    update_balances(root, month)
    return {
        "month": month,
        "target_path": str(target),
        "is_corrupted": True,
        "error": error,
        "action": "repaired",
        "backup_path": str(backup_path),
        "rebuilt_path": str(target),
    }


def all_month_workbooks(root: Path) -> list[Path]:
    if not root.exists():
        return []
    workbooks: list[Path] = []
    for year_dir in sorted(root.iterdir()):
        if not year_dir.is_dir() or not year_dir.name.isdigit() or len(year_dir.name) != 4:
            continue
        for path in year_dir.glob("*.xlsx"):
            if not path.is_file():
                continue
            match = MONTH_WORKBOOK_RE.match(path.name)
            if not match:
                continue
            if match.group("year") != year_dir.name:
                continue
            workbooks.append(path)
    return sorted(workbooks)


def load_transaction_refs(root: Path) -> list[TransactionRef]:
    refs = []
    for workbook_path in all_month_workbooks(root):
        month = workbook_path.stem
        try:
            wb = load_workbook(workbook_path)
        except Exception:
            continue
        for row_index, row in indexed_rows_as_dicts(wb["transactions"]):
            refs.append(TransactionRef(workbook_path, month, row_index, row))
    return refs


def parse_optional_date(raw: str | None) -> date | None:
    return parse_date(raw) if raw else None


def preflight(root: Path, month: str | None) -> dict[str, object]:
    issues: list[dict[str, str]] = []
    ignored_files: list[str] = []
    config = config_path(root)
    if not config.exists():
        issues.append({"code": "missing_config", "path": str(config), "message": "config workbook does not exist"})
    else:
        try:
            wb = load_workbook(config)
            missing = [sheet for sheet in CONFIG_SHEETS if sheet not in wb.sheetnames]
            if missing:
                issues.append({"code": "config_missing_sheets", "path": str(config), "message": f"missing sheets: {','.join(missing)}"})
        except Exception as exc:
            issues.append({"code": "invalid_config", "path": str(config), "message": f"{type(exc).__name__}: {exc}"})

    if month:
        target = month_path(root, month)
        if not target.exists():
            issues.append({"code": "missing_month_workbook", "path": str(target), "message": "month workbook does not exist"})
        else:
            try:
                load_workbook(target)
            except Exception as exc:
                issues.append({"code": "invalid_month_workbook", "path": str(target), "message": f"{type(exc).__name__}: {exc}"})
    elif root.exists():
        for year_dir in sorted(root.iterdir()):
            if not year_dir.is_dir() or not year_dir.name.isdigit() or len(year_dir.name) != 4:
                continue
            for path in sorted(year_dir.glob("*.xlsx")):
                match = MONTH_WORKBOOK_RE.match(path.name)
                if not match or match.group("year") != year_dir.name:
                    ignored_files.append(str(path))
                    continue
                try:
                    load_workbook(path)
                except Exception as exc:
                    issues.append({"code": "invalid_month_workbook", "path": str(path), "message": f"{type(exc).__name__}: {exc}"})

    return {
        "ok": len(issues) == 0,
        "root": str(root),
        "month": month or "",
        "config_exists": config.exists(),
        "issues": issues,
        "ignored_files": ignored_files,
    }


def tags_to_csv(raw: object) -> str:
    if raw is None:
        return ""
    if isinstance(raw, str):
        return normalize_tags(split_tags(raw))
    if isinstance(raw, list):
        return normalize_tags([str(item) for item in raw])
    return normalize_tags(split_tags(str(raw)))


def load_batch_entries(file_path: Path) -> list[dict[str, object]]:
    with file_path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)
    if isinstance(payload, dict):
        payload = payload.get("transactions", [])
    if not isinstance(payload, list):
        raise ValueError("batch payload must be a list or an object with transactions list")
    records: list[dict[str, object]] = []
    for item in payload:
        if not isinstance(item, dict):
            raise ValueError("each batch record must be an object")
        records.append(item)
    return records


def batch_ingest(root: Path, file_path: Path, month: str, dry_run: bool) -> dict[str, object]:
    entries = load_batch_entries(file_path)
    planned: list[dict[str, object]] = []
    for index, item in enumerate(entries, start=1):
        direction = str(item.get("direction", "")).strip()
        if direction not in {"expense", "income", "transfer"}:
            raise ValueError(f"batch row {index}: invalid direction")
        raw_date = str(item.get("date", "")).strip()
        if not raw_date:
            raise ValueError(f"batch row {index}: missing date")
        parsed_day = parse_date(raw_date)
        if month and month_string(parsed_day) != month:
            raise ValueError(f"batch row {index}: date {raw_date} is outside --month {month}")
        amount = parse_money(item.get("amount", ""))
        tags = tags_to_csv(item.get("tags", ""))
        note = str(item.get("note", "")).strip()
        if direction == "transfer":
            from_account = validate_account(root, str(item.get("from_account", "")).strip())
            to_account = validate_account(root, str(item.get("to_account", "")).strip())
            planned.append({
                "direction": direction,
                "date": raw_date,
                "amount": str(amount),
                "from_account": from_account,
                "to_account": to_account,
                "tags": tags,
                "note": note,
            })
            continue
        account = validate_account(root, str(item.get("account", "")).strip())
        planned.append({
            "direction": direction,
            "date": raw_date,
            "amount": str(amount),
            "account": account,
            "level1": str(item.get("level1", "")).strip(),
            "level2": str(item.get("level2", "")).strip(),
            "tags": tags,
            "note": note,
            "counterparty": str(item.get("counterparty", "")).strip(),
            "merchant": str(item.get("merchant", "")).strip(),
            "subscription_name": str(item.get("subscription_name", "")).strip(),
            "subscription_cycle": str(item.get("subscription_cycle", "")).strip(),
            "subscription_start_date": str(item.get("subscription_start_date", "")).strip(),
            "subscription_renewal_date": str(item.get("subscription_renewal_date", "")).strip(),
            "subscription_end_date": str(item.get("subscription_end_date", "")).strip(),
        })

    if dry_run:
        return {
            "total_records": len(planned),
            "imported_count": 0,
            "dry_run": True,
            "records": planned,
        }

    results: list[dict[str, str]] = []
    for item in planned:
        if item["direction"] == "transfer":
            results.append(add_transfer(
                root,
                str(item["date"]),
                parse_money(str(item["amount"])),
                str(item["from_account"]),
                str(item["to_account"]),
                str(item["tags"]),
                str(item["note"]),
            ))
            continue
        results.append(add_transaction(
            root,
            str(item["date"]),
            str(item["direction"]),
            parse_money(str(item["amount"])),
            str(item["account"]),
            str(item["level1"]),
            str(item["level2"]),
            str(item["tags"]),
            str(item["note"]),
            str(item["counterparty"]),
            str(item["merchant"]),
            "",
            str(item["subscription_name"]),
            str(item["subscription_cycle"]),
            str(item["subscription_start_date"]),
            str(item["subscription_renewal_date"]),
            str(item["subscription_end_date"]),
        ))
    return {
        "total_records": len(planned),
        "imported_count": len(results),
        "dry_run": False,
        "transactions": results,
    }


def add_filter_arguments(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("--id", action="append", default=[])
    parser.add_argument("--month", action="append", default=[])
    parser.add_argument("--date-from")
    parser.add_argument("--date-to")
    parser.add_argument("--direction", choices=["expense", "income", "transfer"])
    parser.add_argument("--account")
    parser.add_argument("--from-account")
    parser.add_argument("--to-account")
    parser.add_argument("--level1")
    parser.add_argument("--level2")
    parser.add_argument("--tag", action="append", default=[])
    parser.add_argument("--note-contains", default="")
    parser.add_argument("--merchant-contains", default="")
    parser.add_argument("--counterparty-contains", default="")
    parser.add_argument("--text-contains", default="")


def has_filters(args: argparse.Namespace) -> bool:
    return any([
        args.id,
        args.month,
        args.date_from,
        args.date_to,
        args.direction,
        args.account,
        args.from_account,
        args.to_account,
        args.level1,
        args.level2,
        args.tag,
        args.note_contains,
        args.merchant_contains,
        args.counterparty_contains,
        args.text_contains,
    ])


def row_matches_filters(row: dict[str, str], args: argparse.Namespace) -> bool:
    if args.id and row.get("transaction_id") not in set(args.id):
        return False
    if args.month and month_string(parse_date(row["date"])) not in set(args.month):
        return False
    date_from = parse_optional_date(args.date_from)
    if date_from and parse_date(row["date"]) < date_from:
        return False
    date_to = parse_optional_date(args.date_to)
    if date_to and parse_date(row["date"]) > date_to:
        return False
    if args.direction and row.get("direction") != args.direction:
        return False
    if args.account and row.get("account") != args.account:
        return False
    if args.from_account and row.get("from_account") != args.from_account:
        return False
    if args.to_account and row.get("to_account") != args.to_account:
        return False
    if args.level1 and row.get("level1") != args.level1:
        return False
    if args.level2 and row.get("level2") != args.level2:
        return False
    row_tags = set(split_tags(row.get("tags", "")))
    if args.tag and not set(args.tag).issubset(row_tags):
        return False
    if args.note_contains and args.note_contains.lower() not in row.get("note", "").lower():
        return False
    if args.merchant_contains and args.merchant_contains.lower() not in row.get("merchant", "").lower():
        return False
    if args.counterparty_contains and args.counterparty_contains.lower() not in row.get("counterparty", "").lower():
        return False
    if args.text_contains:
        haystack = combined_text([row.get("note", ""), row.get("merchant", ""), row.get("counterparty", ""), row.get("tags", ""), row.get("level1", ""), row.get("level2", "")])
        if args.text_contains.lower() not in haystack:
            return False
    return True


def sort_key_for_ref(ref: TransactionRef, sort_by: str) -> tuple:
    amount = parse_money(ref.data["amount"])
    if sort_by == "date-asc":
        return (ref.data["date"], ref.data["transaction_id"])
    if sort_by == "amount-asc":
        return (amount, ref.data["date"], ref.data["transaction_id"])
    if sort_by == "amount-desc":
        return (-amount, ref.data["date"], ref.data["transaction_id"])
    return (ref.data["date"], ref.data["transaction_id"])


def query_transactions(root: Path, args: argparse.Namespace) -> list[TransactionRef]:
    if not has_filters(args):
        raise ValueError("query requires at least one filter")
    matches = [ref for ref in load_transaction_refs(root) if row_matches_filters(ref.data, args)]
    reverse = args.sort in {"date-desc", "amount-desc"}
    matches.sort(key=lambda ref: sort_key_for_ref(ref, args.sort), reverse=reverse)
    if args.limit is not None:
        return matches[:args.limit]
    return matches


def serialize_transaction(ref: TransactionRef) -> dict[str, str]:
    row = dict(ref.data)
    row["month"] = ref.month
    return row


def run_query(root: Path, args: argparse.Namespace) -> dict[str, object]:
    matches = query_transactions(root, args)
    return {"count": len(matches), "transactions": [serialize_transaction(ref) for ref in matches]}


def group_keys(row: dict[str, str], group_by: str) -> list[str]:
    if group_by == "none":
        return ["all"]
    if group_by == "month":
        return [month_string(parse_date(row["date"]))]
    if group_by == "date":
        return [row["date"]]
    if group_by == "account":
        return [row.get("account") or row.get("from_account") or row.get("to_account") or "未指定账户"]
    if group_by == "level1":
        return [row.get("level1") or "未分类"]
    if group_by == "level2":
        return [row.get("level2") or "未分类"]
    if group_by == "category":
        if row.get("level1") and row.get("level2"):
            return [f"{row['level1']}/{row['level2']}"]
        return ["未分类"]
    if group_by == "tag":
        tags = split_tags(row.get("tags", ""))
        return tags or ["未打标签"]
    if group_by == "merchant":
        return [row.get("merchant") or "未填写商户"]
    if group_by == "counterparty":
        return [row.get("counterparty") or "未填写对方"]
    if group_by == "direction":
        return [row.get("direction", "")]
    raise ValueError(f"unsupported group_by: {group_by}")


def run_stats(root: Path, args: argparse.Namespace) -> dict[str, object]:
    matches = query_transactions(root, args)
    total = Decimal("0.00")
    groups: dict[str, dict[str, Decimal | int]] = {}
    for ref in matches:
        amount = parse_money(ref.data["amount"])
        total += amount
        for key in group_keys(ref.data, args.group_by):
            bucket = groups.setdefault(key, {"count": 0, "total_amount": Decimal("0.00")})
            bucket["count"] = int(bucket["count"]) + 1
            bucket["total_amount"] = Decimal(bucket["total_amount"]) + amount
    ordered_groups = sorted(
        groups.items(),
        key=lambda item: (item[0] if args.group_by in {"month", "date"} else str(-Decimal(item[1]["total_amount"]))),
    )
    if args.group_by not in {"month", "date"}:
        ordered_groups = sorted(groups.items(), key=lambda item: (-Decimal(item[1]["total_amount"]), item[0]))
    return {
        "count": len(matches),
        "total_amount": str(total.quantize(MONEY_QUANT)),
        "group_by": args.group_by,
        "groups": [
            {
                "group": key,
                "count": int(value["count"]),
                "total_amount": str(Decimal(value["total_amount"]).quantize(MONEY_QUANT)),
            }
            for key, value in ordered_groups
        ],
    }


def update_tags(existing: str, set_tags: str, add_tags: str, remove_tags: str) -> str:
    if set_tags:
        tags = split_tags(set_tags)
    else:
        tags = split_tags(existing)
        tags.extend(split_tags(add_tags))
        remove = set(split_tags(remove_tags))
        tags = [tag for tag in tags if tag not in remove]
    return normalize_tags(tags)


def apply_update(root: Path, row: dict[str, str], args: argparse.Namespace) -> dict[str, str]:
    updated = dict(row)
    if args.set_date:
        updated["date"] = args.set_date
    if args.set_amount:
        updated["amount"] = str(parse_money(args.set_amount))
    if row["direction"] == "transfer":
        if args.set_account:
            raise ValueError("transfer transaction does not support --set-account")
        if args.set_from_account:
            updated["from_account"] = validate_account(root, args.set_from_account)
        if args.set_to_account:
            updated["to_account"] = validate_account(root, args.set_to_account)
    else:
        if args.set_account:
            updated["account"] = validate_account(root, args.set_account)
    if args.set_level1 or args.set_level2:
        next_level1 = args.set_level1 if args.set_level1 is not None else updated["level1"]
        next_level2 = args.set_level2 if args.set_level2 is not None else updated["level2"]
        validate_category(root, updated["direction"], next_level1, next_level2)
        updated["level1"] = next_level1
        updated["level2"] = next_level2
        updated["category_source"] = "manual"
    if args.set_tags or args.add_tags or args.remove_tags:
        updated["tags"] = update_tags(updated["tags"], args.set_tags, args.add_tags, args.remove_tags)
    if args.set_note is not None:
        updated["note"] = args.set_note
    if args.set_merchant is not None:
        updated["merchant"] = args.set_merchant
    if args.set_counterparty is not None:
        updated["counterparty"] = args.set_counterparty
    if args.set_subscription_name is not None:
        updated["subscription_name"] = args.set_subscription_name
    if args.set_subscription_cycle is not None:
        updated["subscription_cycle"] = args.set_subscription_cycle
    if args.set_subscription_start_date is not None:
        updated["subscription_start_date"] = args.set_subscription_start_date
    if args.set_subscription_renewal_date is not None:
        updated["subscription_renewal_date"] = args.set_subscription_renewal_date
    if args.set_subscription_end_date is not None:
        updated["subscription_end_date"] = args.set_subscription_end_date
    return updated


def run_update(root: Path, args: argparse.Namespace) -> dict[str, object]:
    if not has_filters(args):
        raise ValueError("update requires selectors")
    if not any([
        args.set_date,
        args.set_amount,
        args.set_account,
        args.set_from_account,
        args.set_to_account,
        args.set_level1 is not None,
        args.set_level2 is not None,
        args.set_tags,
        args.add_tags,
        args.remove_tags,
        args.set_note is not None,
        args.set_merchant is not None,
        args.set_counterparty is not None,
        args.set_subscription_name is not None,
        args.set_subscription_cycle is not None,
        args.set_subscription_start_date is not None,
        args.set_subscription_renewal_date is not None,
        args.set_subscription_end_date is not None,
    ]):
        raise ValueError("update requires at least one mutation")
    matches = query_transactions(root, args)
    if not matches:
        raise ValueError("no transactions matched the selectors")
    if len(matches) > 1 and not (args.allow_multiple or len(args.id) > 1):
        raise ValueError("multiple transactions matched; rerun with --allow-multiple or explicit --id selectors")

    updated_records = []
    touched_months = set()
    deletions: dict[Path, list[int]] = defaultdict(list)
    for ref in matches:
        updated = apply_update(root, ref.data, args)
        updated_records.append((ref, updated))
        deletions[ref.workbook_path].append(ref.row_index)
        touched_months.add(ref.month)
        touched_months.add(month_string(parse_date(updated["date"])))

    for workbook_path, row_indices in deletions.items():
        wb = load_workbook(workbook_path)
        sheet = wb["transactions"]
        for row_index in sorted(row_indices, reverse=True):
            sheet.delete_rows(row_index)
        save_workbook_atomic(wb, workbook_path)

    for _, updated in updated_records:
        target_month = month_string(parse_date(updated["date"]))
        ensure_month_workbook(root, target_month)
        wb = load_workbook(month_path(root, target_month))
        append_row(wb["transactions"], MONTHLY_SHEETS["transactions"], updated)
        save_workbook_atomic(wb, month_path(root, target_month))

    for month in sorted(touched_months):
        update_balances(root, month)
    rebuild_subscriptions(root)
    return {
        "updated_count": len(updated_records),
        "transactions": [dict(item[1], month=month_string(parse_date(item[1]["date"]))) for item in updated_records],
    }


def run_subscription_report(root: Path, args: argparse.Namespace) -> dict[str, object]:
    rebuild_subscriptions(root)
    rows = load_config_rows(root, "subscriptions")
    today = date.today()
    filtered = []
    for row in rows:
        if args.status and row.get("status") != args.status:
            continue
        if args.due_within is not None:
            renewal = row.get("renewal_date")
            if not renewal:
                continue
            remaining = (parse_date(renewal) - today).days
            if remaining < 0 or remaining > args.due_within:
                continue
        filtered.append(row)
    filtered.sort(key=lambda row: (row.get("renewal_date") or "9999-12-31", row.get("name", "")))
    return {"count": len(filtered), "subscriptions": filtered}


def format_result(result: object) -> str:
    return json.dumps(result, ensure_ascii=False, indent=2)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Excel bookkeeping helper")
    sub = parser.add_subparsers(dest="command", required=True)
    mutates = "[mutates]"
    read_request_may_mutate = "[read-request-may-mutate]"
    read_only = "[read-only]"

    boot = sub.add_parser("bootstrap", help=f"{mutates} initialize config and month workbook")
    boot.add_argument("--root", required=True)
    boot.add_argument("--month")

    add = sub.add_parser("add", help=f"{mutates} add one income/expense transaction")
    add.add_argument("--root", required=True)
    add.add_argument("--date", required=True)
    add.add_argument("--direction", choices=["expense", "income"], required=True)
    add.add_argument("--amount", required=True)
    add.add_argument("--account", required=True)
    add.add_argument("--level1", default="")
    add.add_argument("--level2", default="")
    add.add_argument("--tags", default="")
    add.add_argument("--note", default="")
    add.add_argument("--counterparty", default="")
    add.add_argument("--merchant", default="")
    add.add_argument("--subscription-name", default="")
    add.add_argument("--subscription-cycle", choices=["", "monthly", "yearly"], default="")
    add.add_argument("--subscription-start-date", default="")
    add.add_argument("--subscription-renewal-date", default="")
    add.add_argument("--subscription-end-date", default="")

    transfer = sub.add_parser("transfer", help=f"{mutates} add one transfer transaction")
    transfer.add_argument("--root", required=True)
    transfer.add_argument("--date", required=True)
    transfer.add_argument("--amount", required=True)
    transfer.add_argument("--from-account", required=True)
    transfer.add_argument("--to-account", required=True)
    transfer.add_argument("--tags", default="")
    transfer.add_argument("--note", default="")

    for command in ("set-opening", "set-closing"):
        balance = sub.add_parser(command, help=f"{mutates} update monthly balance fields")
        balance.add_argument("--root", required=True)
        balance.add_argument("--month", required=True)
        balance.add_argument("--account", required=True)
        balance.add_argument("--amount", required=True)

    recurring = sub.add_parser("add-recurring", help=f"{mutates} add recurring rule")
    recurring.add_argument("--root", required=True)
    recurring.add_argument("--name", required=True)
    recurring.add_argument("--direction", choices=["expense", "income", "transfer"], required=True)
    recurring.add_argument("--amount", required=True)
    recurring.add_argument("--account", default="")
    recurring.add_argument("--from-account", default="")
    recurring.add_argument("--to-account", default="")
    recurring.add_argument("--level1", default="")
    recurring.add_argument("--level2", default="")
    recurring.add_argument("--tags", default="")
    recurring.add_argument("--note", default="")
    recurring.add_argument("--counterparty", default="")
    recurring.add_argument("--merchant", default="")
    recurring.add_argument("--frequency", choices=["monthly", "yearly"], required=True)
    recurring.add_argument("--interval", type=int, default=1)
    recurring.add_argument("--month", type=int, default=0)
    recurring.add_argument("--day", type=int, default=1)
    recurring.add_argument("--start-date", required=True)
    recurring.add_argument("--end-date", default="")

    apply_cmd = sub.add_parser("apply-recurring", help=f"{mutates} materialize recurring rules into a month")
    apply_cmd.add_argument("--root", required=True)
    apply_cmd.add_argument("--month", required=True)

    limit_cmd = sub.add_parser("set-limit", help=f"{mutates} set annual tag limit")
    limit_cmd.add_argument("--root", required=True)
    limit_cmd.add_argument("--year", type=int, required=True)
    limit_cmd.add_argument("--tag", required=True)
    limit_cmd.add_argument("--amount", required=True)
    limit_cmd.add_argument("--note", default="")

    limit_report_cmd = sub.add_parser("limit-report", help=f"{read_only} report annual tag usage")
    limit_report_cmd.add_argument("--root", required=True)
    limit_report_cmd.add_argument("--year", type=int, required=True)

    list_accounts_cmd = sub.add_parser("list-accounts", help=f"{read_only} list configured accounts and aliases")
    list_accounts_cmd.add_argument("--root", required=True)
    list_accounts_cmd.add_argument("--include-inactive", action="store_true")

    add_account_cmd = sub.add_parser("add-account", help=f"{mutates} add or update an account and optional aliases")
    add_account_cmd.add_argument("--root", required=True)
    add_account_cmd.add_argument("--account", required=True)
    add_account_cmd.add_argument("--account-type", default="wallet")
    add_account_cmd.add_argument("--currency", default="CNY")
    add_account_cmd.add_argument("--opening-hint", default="0")
    add_account_cmd.add_argument("--note", default="")
    add_account_cmd.add_argument("--inactive", action="store_true")
    add_account_cmd.add_argument("--alias", action="append", default=[])

    month_report_cmd = sub.add_parser("month-report", help=f"{read_request_may_mutate} generate month report")
    month_report_cmd.add_argument("--root", required=True)
    month_report_cmd.add_argument("--month", required=True)

    preflight_cmd = sub.add_parser("preflight", help=f"{read_only} check config and workbook health")
    preflight_cmd.add_argument("--root", required=True)
    preflight_cmd.add_argument("--month")

    query_cmd = sub.add_parser("query", help=f"{read_request_may_mutate} filter transactions")
    query_cmd.add_argument("--root", required=True)
    add_filter_arguments(query_cmd)
    query_cmd.add_argument("--sort", choices=["date-desc", "date-asc", "amount-desc", "amount-asc"], default="date-desc")
    query_cmd.add_argument("--limit", type=int)

    stats_cmd = sub.add_parser("stats", help=f"{read_request_may_mutate} aggregate transactions")
    stats_cmd.add_argument("--root", required=True)
    add_filter_arguments(stats_cmd)
    stats_cmd.add_argument("--group-by", choices=["none", "month", "date", "account", "level1", "level2", "category", "tag", "merchant", "counterparty", "direction"], default="none")
    stats_cmd.add_argument("--sort", choices=["date-desc", "date-asc", "amount-desc", "amount-asc"], default="date-desc")
    stats_cmd.add_argument("--limit", type=int)

    update_cmd = sub.add_parser("update", help=f"{mutates} update matched transactions")
    update_cmd.add_argument("--root", required=True)
    add_filter_arguments(update_cmd)
    update_cmd.add_argument("--sort", choices=["date-desc", "date-asc", "amount-desc", "amount-asc"], default="date-desc")
    update_cmd.add_argument("--limit", type=int)
    update_cmd.add_argument("--allow-multiple", action="store_true")
    update_cmd.add_argument("--set-date")
    update_cmd.add_argument("--set-amount")
    update_cmd.add_argument("--set-account")
    update_cmd.add_argument("--set-from-account")
    update_cmd.add_argument("--set-to-account")
    update_cmd.add_argument("--set-level1")
    update_cmd.add_argument("--set-level2")
    update_cmd.add_argument("--set-tags", default="")
    update_cmd.add_argument("--add-tags", default="")
    update_cmd.add_argument("--remove-tags", default="")
    update_cmd.add_argument("--set-note")
    update_cmd.add_argument("--set-merchant")
    update_cmd.add_argument("--set-counterparty")
    update_cmd.add_argument("--set-subscription-name")
    update_cmd.add_argument("--set-subscription-cycle")
    update_cmd.add_argument("--set-subscription-start-date")
    update_cmd.add_argument("--set-subscription-renewal-date")
    update_cmd.add_argument("--set-subscription-end-date")

    subscriptions_cmd = sub.add_parser("subscription-report", help=f"{read_request_may_mutate} rebuild and report subscriptions")
    subscriptions_cmd.add_argument("--root", required=True)
    subscriptions_cmd.add_argument("--status", choices=["active", "due", "expired"])
    subscriptions_cmd.add_argument("--due-within", type=int)

    repair_cmd = sub.add_parser("repair-month", help=f"{mutates} dry-run or repair one corrupted monthly workbook")
    repair_cmd.add_argument("--root", required=True)
    repair_cmd.add_argument("--month", required=True)
    repair_cmd.add_argument("--execute", action="store_true")

    batch_ingest_cmd = sub.add_parser("batch-ingest", help=f"{mutates} import batch transactions from JSON")
    batch_ingest_cmd.add_argument("--root", required=True)
    batch_ingest_cmd.add_argument("--file", required=True)
    batch_ingest_cmd.add_argument("--month", default="")
    batch_ingest_cmd.add_argument("--dry-run", action="store_true")

    return parser


def main() -> None:
    args = build_parser().parse_args()
    root = Path(args.root).expanduser().resolve()
    with ledger_lock(root):
        if args.command == "bootstrap":
            result = bootstrap(root, args.month)
        elif args.command == "preflight":
            result = preflight(root, args.month)
        elif args.command == "repair-month":
            result = repair_month(root, args.month, args.execute)
        else:
            if args.command not in {"add-account", "list-accounts"} and not config_path(root).exists():
                raise SystemExit("run bootstrap first")
            if args.command == "add":
                result = add_transaction(
                    root,
                    args.date,
                    args.direction,
                    parse_money(args.amount),
                    args.account,
                    args.level1,
                    args.level2,
                    args.tags,
                    args.note,
                    args.counterparty,
                    args.merchant,
                    "",
                    args.subscription_name,
                    args.subscription_cycle,
                    args.subscription_start_date,
                    args.subscription_renewal_date,
                    args.subscription_end_date,
                )
            elif args.command == "transfer":
                result = add_transfer(root, args.date, parse_money(args.amount), args.from_account, args.to_account, args.tags, args.note)
            elif args.command == "set-opening":
                result = set_balance(root, args.month, args.account, parse_money(args.amount), "opening_balance")
            elif args.command == "set-closing":
                result = set_balance(root, args.month, args.account, parse_money(args.amount), "closing_balance")
            elif args.command == "add-recurring":
                result = add_recurring(
                    root,
                    args.name,
                    args.direction,
                    parse_money(args.amount),
                    args.account,
                    args.from_account,
                    args.to_account,
                    args.level1,
                    args.level2,
                    args.tags,
                    args.note,
                    args.counterparty,
                    args.merchant,
                    args.frequency,
                    args.interval,
                    args.month,
                    args.day,
                    args.start_date,
                    args.end_date,
                )
            elif args.command == "apply-recurring":
                result = apply_recurring(root, args.month)
            elif args.command == "set-limit":
                result = upsert_limit(root, args.year, args.tag, parse_money(args.amount), args.note)
            elif args.command == "limit-report":
                result = collect_limit_usage(root, args.year)
            elif args.command == "list-accounts":
                result = list_accounts(root, args.include_inactive)
            elif args.command == "add-account":
                result = upsert_account(
                    root,
                    args.account,
                    args.account_type,
                    args.currency,
                    args.opening_hint,
                    not args.inactive,
                    args.note,
                    args.alias,
                )
            elif args.command == "month-report":
                result = month_report(root, args.month)
            elif args.command == "query":
                result = run_query(root, args)
            elif args.command == "stats":
                result = run_stats(root, args)
            elif args.command == "update":
                result = run_update(root, args)
            elif args.command == "subscription-report":
                result = run_subscription_report(root, args)
            elif args.command == "batch-ingest":
                result = batch_ingest(root, Path(args.file).expanduser().resolve(), args.month, args.dry_run)
            else:
                raise SystemExit(f"unsupported command: {args.command}")
    print(format_result(result))


if __name__ == "__main__":
    main()
